"""LangChain-based document parser adapter."""

from __future__ import annotations

import logging
from pathlib import Path

from misaka.services.knowledge.rag.abstractions import DocumentParser, ParsedDocument

logger = logging.getLogger(__name__)


class LCDocumentParser(DocumentParser):
    """Parse documents via LangChain DocumentLoaders.

    Loader mapping:
        txt / markdown → TextLoader
        docx           → Docx2txtLoader
        pdf            → PyMuPDFLoader
        xlsx           → custom openpyxl loader
    """

    _SUPPORTED: list[str] = ["txt", "markdown", "docx", "xlsx", "pdf"]

    async def parse(self, file_path: str, file_type: str) -> ParsedDocument:
        if file_type not in self._SUPPORTED:
            raise ValueError(
                f"Unsupported file type '{file_type}'. "
                f"Supported: {self._SUPPORTED}"
            )

        loader = self._select_loader(file_path, file_type)
        lc_docs = loader.load()

        if not lc_docs:
            return ParsedDocument(text="", metadata={"source": file_path})

        full_text = "\n\n".join(doc.page_content for doc in lc_docs)
        metadata = {**lc_docs[0].metadata, "source": file_path, "file_type": file_type}
        page_breaks = self._extract_page_breaks(lc_docs)

        return ParsedDocument(
            text=full_text,
            metadata=metadata,
            page_breaks=page_breaks,
        )

    def supported_types(self) -> list[str]:
        return list(self._SUPPORTED)

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _select_loader(self, file_path: str, file_type: str):  # noqa: ANN202
        """Return the appropriate LangChain loader instance."""
        if file_type in ("txt", "markdown"):
            return self._text_loader(file_path)
        if file_type == "docx":
            return self._docx_loader(file_path)
        if file_type == "pdf":
            return self._pdf_loader(file_path)
        if file_type == "xlsx":
            return self._xlsx_loader(file_path)
        raise ValueError(f"No loader available for '{file_type}'")

    @staticmethod
    def _text_loader(file_path: str):  # noqa: ANN205
        from langchain_community.document_loaders import TextLoader

        return TextLoader(file_path, encoding="utf-8", autodetect_encoding=True)

    @staticmethod
    def _docx_loader(file_path: str):  # noqa: ANN205
        from langchain_community.document_loaders import Docx2txtLoader

        return Docx2txtLoader(file_path)

    @staticmethod
    def _pdf_loader(file_path: str):  # noqa: ANN205
        from langchain_community.document_loaders import PyMuPDFLoader

        return PyMuPDFLoader(file_path)

    @staticmethod
    def _xlsx_loader(file_path: str):  # noqa: ANN205
        """Build a lightweight openpyxl-based loader (returns LangChain Documents)."""
        return _OpenpyxlLoader(file_path)

    @staticmethod
    def _extract_page_breaks(lc_docs: list) -> list[int]:
        """Collect page numbers from LangChain document metadata."""
        breaks: list[int] = []
        for doc in lc_docs:
            page = doc.metadata.get("page")
            if page is not None:
                breaks.append(int(page))
        return breaks


# ---------------------------------------------------------------------------
# Minimal openpyxl-based loader (not provided by LangChain out of the box)
# ---------------------------------------------------------------------------

class _OpenpyxlLoader:
    """Load an Excel workbook into LangChain-compatible Document objects.

    Each worksheet becomes one Document whose ``page_content`` is a
    tab-separated text representation of the rows.
    """

    def __init__(self, file_path: str) -> None:
        self._file_path = file_path

    def load(self) -> list:
        from langchain_core.documents import Document
        import openpyxl

        path = Path(self._file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {self._file_path}")

        wb = openpyxl.load_workbook(self._file_path, read_only=True, data_only=True)
        documents: list[Document] = []

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue

                lines: list[str] = []
                for row in rows:
                    line = "\t".join(str(cell) if cell is not None else "" for cell in row)
                    lines.append(line)

                documents.append(Document(
                    page_content="\n".join(lines),
                    metadata={
                        "source": self._file_path,
                        "sheet": sheet_name,
                        "row_count": len(rows),
                    },
                ))
        finally:
            wb.close()

        return documents
